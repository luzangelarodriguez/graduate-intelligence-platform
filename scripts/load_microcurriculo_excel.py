#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
scripts/load_microcurriculo_excel.py
-------------------------------------
Carga microcurrículos desde archivos .xlsx en storage/test_microcurriculos/
a la base de datos Railway (tabla microcurriculos + microcurriculo_skills).

Estructura esperada del Excel:
  - Hoja: 'Matríz de RA' (o primera hoja si no existe)
  - Desde fila 13 (índice 12):
      columna 8  (índice 7)  → nombre asignatura
      columna 10 (índice 9)  → resultado de aprendizaje (RA)
      columna 11 (índice 10) → créditos

Usage:
    python scripts/load_microcurriculo_excel.py --preview
    python scripts/load_microcurriculo_excel.py --execute
    python scripts/load_microcurriculo_excel.py --file path/to/archivo.xlsx --preview
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any

# ── paths ──────────────────────────────────────────────────────────────────────
REPO_ROOT   = Path(__file__).resolve().parent.parent
STORAGE_DIR = REPO_ROOT / "storage" / "test_microcurriculos"
ENV_LOCAL   = REPO_ROOT / ".env.local"

# ── hard-coded program mapping ─────────────────────────────────────────────────
# folder name fragment (normalizado) → especializacion_id
EXPLICIT_ESP_MAP: dict[str, int] = {
    "neuropsicolog": 20,
    "neuropsicologia": 20,
    "psicolog": 20,
    "criminolog": 108,
    "victimolog": 108,
    "visual analytics": 94,
    "big data": 94,
    "inteligencia artificial": 92,
    "machine learning": 92,
    "administracion de empresas": 82,
    "alta gerencia": 82,
}

# ── skill keywords para neuropsicología ───────────────────────────────────────
NEURO_KEYWORDS: list[str] = [
    "neuropsicolog",
    "evaluacion",
    "intervencion",
    "funciones ejecutivas",
    "memoria",
    "atencion",
    "percepcion",
    "lenguaje",
    "lectura",
    "escritura",
    "aprendizaje",
    "trastorno",
    "discapacidad",
    "inclusion",
    "cognitiv",
    "neuroling",
    "inteligencia",
    "desarrollo",
    "educacion",
    "funciones superiores",
    "rehabilitacion",
    "diagnostico",
    "diversidad",
    "orientacion",
]

# ── texto normalización ────────────────────────────────────────────────────────
def normalize(text: str) -> str:
    """Lowercase + strip diacritics."""
    nfkd = unicodedata.normalize("NFKD", text.lower())
    return "".join(c for c in nfkd if not unicodedata.combining(c)).strip()


def clean(text: Any) -> str:
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text).strip())


# ── skill extraction desde texto de RAs ───────────────────────────────────────
_RE_TECH = re.compile(
    r"\b(Python|R\b|SQL|NoSQL|SPSS|Excel|Power\s*BI|Tableau|"
    r"Machine\s*Learning|Deep\s*Learning|NLP|"
    r"Visualizaci[oó]n\s*de\s*datos|An[aá]lisis\s*de\s*datos)\b",
    re.IGNORECASE,
)
_RE_TRANSVERSAL = re.compile(
    r"\b(Comunicaci[oó]n|Liderazgo|Trabajo\s+en\s+equipo|"
    r"Pensamiento\s+cr[ií]tico|Toma\s+de\s+decisiones|"
    r"Innovaci[oó]n|[EÉ]tica|Resoluci[oó]n\s+de\s+problemas)\b",
    re.IGNORECASE,
)


def extract_skills_from_text(text: str, domain_keywords: list[str]) -> dict[str, list[str]]:
    """Extract skills: domain keywords + regex tech/transversal."""
    seen: set[str] = set()
    buckets: dict[str, list[str]] = {
        "competencia_dominio": [],
        "tecnologia": [],
        "skill_transversal": [],
    }
    norm_text = normalize(text)

    # Domain-specific keywords
    for kw in domain_keywords:
        if kw in norm_text and kw not in seen:
            seen.add(kw)
            buckets["competencia_dominio"].append(kw.replace("_", " ").title())

    # Tech tools
    for m in _RE_TECH.finditer(text):
        key = normalize(m.group(0))
        if key not in seen:
            seen.add(key)
            buckets["tecnologia"].append(m.group(0).strip())

    # Soft skills
    for m in _RE_TRANSVERSAL.finditer(text):
        key = normalize(m.group(0))
        if key not in seen:
            seen.add(key)
            buckets["skill_transversal"].append(m.group(0).strip())

    return {k: v for k, v in buckets.items() if v}


# ── infer especialización id ──────────────────────────────────────────────────
def infer_esp_id(folder_name: str, esps: list[dict]) -> int | None:
    norm_folder = normalize(folder_name)

    # 1. Explicit map
    for fragment, eid in EXPLICIT_ESP_MAP.items():
        if fragment in norm_folder:
            return eid

    # 2. Fuzzy match against DB especializaciones
    for e in esps:
        e_n = normalize(e["nombre"])
        if e_n in norm_folder or norm_folder in e_n:
            return e["id"]

    return None


# ── parse one .xlsx file ──────────────────────────────────────────────────────
def parse_excel(path: Path) -> list[dict[str, Any]]:
    """
    Parse microcurriculum matrix from .xlsx.
    Returns list of dicts, one per asignatura.
    """
    try:
        import openpyxl  # type: ignore
    except ImportError:
        print("ERROR: openpyxl no instalado. Ejecutar: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(str(path), data_only=True)

    # Select sheet
    target_sheet = None
    for name in wb.sheetnames:
        if "matri" in normalize(name) and "ra" in normalize(name):
            target_sheet = wb[name]
            break
    if target_sheet is None:
        target_sheet = wb.active
        print(f"  [WARN] Hoja 'Matríz de RA' no encontrada — usando '{target_sheet.title}'")
    else:
        print(f"  [OK] Usando hoja: '{target_sheet.title}'")

    # Detect program name: use folder name if file is inside a subfolder,
    # otherwise fall back to the filename (without extension)
    if path.parent.resolve() == STORAGE_DIR.resolve():
        folder_name = path.stem  # root-level file → use filename
    else:
        folder_name = path.parent.name
    programa = folder_name.strip()

    # Real column layout (1-indexed, verified from the Neuropsicología matrix):
    #   Col  1 = CÓDIGO (CE1, CE2…) — non-empty signals a new asignatura row
    #   Col  9 = ASIGNATURA name
    #   Col 11 = Resultado de Aprendizaje (RA) text
    #   Col 12 = CRÉDITOS
    START_ROW    = 13
    COL_CODIGO   = 1    # new-asignatura sentinel
    COL_ASIG     = 9    # asignatura name
    COL_RA       = 11   # RA text
    COL_CREDITOS = 12   # credits

    asig_data: dict[str, dict[str, Any]] = {}  # asignatura key → {name, ras, creditos}
    current_key  = ""   # normalized key for dict lookup
    current_name = ""   # original display name

    all_rows = list(target_sheet.iter_rows(min_row=START_ROW, values_only=True))
    print(f"  [INFO] {len(all_rows)} filas desde fila {START_ROW}")

    for row in all_rows:
        row = list(row) + [None] * 15   # pad to avoid IndexError

        codigo_raw = clean(row[COL_CODIGO - 1])    # col 1
        asig_raw   = clean(row[COL_ASIG - 1])      # col 9
        ra_raw     = clean(row[COL_RA - 1])         # col 11
        cred_raw   = clean(row[COL_CREDITOS - 1])   # col 12

        # New asignatura: código cell is non-empty AND asignatura cell is non-empty
        if codigo_raw and asig_raw and len(asig_raw) > 3:
            norm_a = normalize(asig_raw)
            skip = {"asignatura", "materia", "nombre", "resultados de aprendizaje de la asignatura"}
            if norm_a not in skip:
                current_key  = norm_a[:60]
                current_name = re.sub(r"\s+", " ", asig_raw.replace("\n", " ")).strip()
                if current_key not in asig_data:
                    cred_val = None
                    try:
                        cred_val = int(float(cred_raw)) if cred_raw else None
                    except (ValueError, TypeError):
                        pass
                    asig_data[current_key] = {
                        "name": current_name,
                        "ras": [],
                        "creditos": cred_val,
                    }

        # Append any RA text to current asignatura (strip leading ENE-RA-XX codes optionally)
        if ra_raw and len(ra_raw) > 10 and current_key:
            # Split on embedded newlines (some cells pack multiple RAs)
            for part in ra_raw.split("\n"):
                part = part.strip()
                if len(part) > 15:
                    asig_data[current_key]["ras"].append(part)

    if not asig_data:
        print(f"  [WARN] No se encontraron asignaturas en {path.name}")
        return []

    doc_hash_base = hashlib.md5(path.read_bytes()).hexdigest()
    results: list[dict[str, Any]] = []

    for key, info in asig_data.items():
        asig_name: str    = info["name"]
        ras: list[str]    = info["ras"]
        creditos: int | None = info["creditos"]
        full_text = "\n".join(ras)
        skills = extract_skills_from_text(full_text, NEURO_KEYWORDS)
        doc_hash = doc_hash_base + f":{key[:30]}"

        results.append({
            "programa": programa,
            "asignatura": asig_name,
            "creditos": creditos,
            "resultados_aprendizaje": ras[:15],
            "contenido_tematico": [],
            "descripcion": "",
            "skills": skills,
            "source_document": path.name,
            "source_path": str(path.relative_to(REPO_ROOT)),
            "document_hash": doc_hash,
            "domain_key": "neuropsychology",
        })

    return results


# ── discover all .xlsx files ──────────────────────────────────────────────────
def discover_excel_files(target_path: Path | None = None) -> list[Path]:
    import glob as _glob
    base = target_path or STORAGE_DIR
    # Files in any subfolder
    found = list(_glob.glob(str(base / "**" / "*.xlsx"), recursive=True))
    # Files directly in the root of STORAGE_DIR
    found += list(_glob.glob(str(base / "*.xlsx")))
    # Deduplicate, exclude Excel temp files (~$...), sort
    seen: set[str] = set()
    result: list[Path] = []
    for p in sorted(found):
        path = Path(p).resolve()
        if path.name.startswith("~$"):
            continue
        key = str(path)
        if key not in seen:
            seen.add(key)
            result.append(path)
    return result


# ── DB helpers ────────────────────────────────────────────────────────────────
def _load_env() -> None:
    if not ENV_LOCAL.exists():
        return
    try:
        from dotenv import load_dotenv  # type: ignore
        load_dotenv(ENV_LOCAL, override=True)
    except ImportError:
        # Manual fallback (no python-dotenv installed)
        for line in ENV_LOCAL.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, _, v = line.partition("=")
                os.environ[k.strip()] = v.strip().strip('"').strip("'")


def get_conn():
    _load_env()
    url = os.environ.get("RAILWAY_DATABASE_URL") or os.environ.get("DATABASE_URL")
    if not url:
        raise RuntimeError(
            "RAILWAY_DATABASE_URL no configurado.\n"
            "Agregar en .env.local:\n"
            "  RAILWAY_DATABASE_URL=postgresql://user:pass@host:port/db"
        )
    import psycopg2  # type: ignore
    conn = psycopg2.connect(url)
    conn.set_client_encoding("UTF8")
    return conn


def fetch_especializaciones(conn) -> list[dict]:
    with conn.cursor() as cur:
        cur.execute("SELECT id, nombre FROM especializaciones ORDER BY id")
        return [{"id": r[0], "nombre": r[1]} for r in cur.fetchall()]


def fetch_existing_hashes(conn) -> set[str]:
    with conn.cursor() as cur:
        cur.execute(
            "SELECT document_hash FROM microcurriculos WHERE document_hash IS NOT NULL"
        )
        return {r[0] for r in cur.fetchall()}


def insert_record(conn, rec: dict, esp_id: int | None) -> int | None:
    clean_text = "\n\n".join(filter(None, [
        rec.get("descripcion", ""),
        "\n".join(rec.get("resultados_aprendizaje", [])),
    ])).strip()

    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO microcurriculos
                (programa, asignatura, semestre, source_document,
                 document_hash, clean_text, detected_domain, lineage,
                 specialization_id, specialization_name)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (document_hash) DO NOTHING
            RETURNING id
            """,
            (
                rec["programa"],
                rec["asignatura"],
                str(rec.get("creditos") or ""),
                rec["source_document"],
                rec["document_hash"],
                clean_text,
                rec["domain_key"],
                json.dumps({
                    "source_path": rec["source_path"],
                    "domain_key": rec["domain_key"],
                }),
                esp_id,
                rec["programa"] if esp_id else None,
            ),
        )
        row = cur.fetchone()
        if row is None:
            # Already exists — return existing id
            cur.execute(
                "SELECT id FROM microcurriculos WHERE document_hash = %s",
                (rec["document_hash"],),
            )
            existing = cur.fetchone()
            return existing[0] if existing else None
        return row[0]


def insert_skills(conn, micro_id: int, rec: dict) -> int:
    rows = 0
    with conn.cursor() as cur:
        for tipo, skills in rec.get("skills", {}).items():
            for skill in skills:
                if not skill or not str(skill).strip():
                    continue
                try:
                    cur.execute(
                        """
                        INSERT INTO microcurriculo_skills
                            (microcurriculo_id, skill_original, skill_normalized,
                             skill_domain, tipo_skill, confidence_score, source_document)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                        ON CONFLICT (microcurriculo_id, skill_normalized, tipo_skill) DO NOTHING
                        """,
                        (
                            micro_id,
                            skill,
                            normalize(skill),
                            rec["domain_key"],
                            tipo,
                            0.75,
                            rec["source_document"],
                        ),
                    )
                    rows += cur.rowcount
                except Exception as exc:
                    pass  # individual skill conflicts are non-fatal
    return rows


# ── preview printer ───────────────────────────────────────────────────────────
def print_preview(records: list[dict], esp_map: dict[str, int | None]) -> None:
    print("\n" + "=" * 70)
    print(f"PREVIEW — {len(records)} asignatura(s) encontradas")
    print("=" * 70)
    for rec in records:
        esp_id = esp_map.get(rec["programa"])
        skill_count = sum(len(v) for v in rec.get("skills", {}).values())
        ra_count = len(rec.get("resultados_aprendizaje", []))
        cred = rec.get("creditos") or "—"
        print(f"\n  📘 {rec['asignatura']}")
        print(f"     Programa  : {rec['programa']}")
        print(f"     ESP ID    : {esp_id or 'NO ENCONTRADO ⚠'}")
        print(f"     Créditos  : {cred}")
        print(f"     RAs       : {ra_count}")
        print(f"     Skills    : {skill_count}")
        if rec.get("resultados_aprendizaje"):
            print(f"     Primer RA : {rec['resultados_aprendizaje'][0][:80]}…")
        if rec.get("skills"):
            for tipo, skl in rec["skills"].items():
                print(f"     [{tipo}] {', '.join(skl[:5])}")
    print()


# ── main ──────────────────────────────────────────────────────────────────────
def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Cargar microcurrículos desde .xlsx a Railway DB.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument("--preview",  action="store_true", help="Solo mostrar qué se insertaría (sin escribir en DB)")
    p.add_argument("--execute",  action="store_true", help="Insertar en DB (requiere confirmación)")
    p.add_argument("--file",     type=Path, default=None, metavar="PATH", help="Ruta específica a un .xlsx")
    p.add_argument("--yes",      action="store_true", help="Omitir confirmación interactiva")
    return p.parse_args()


def main() -> None:
    args = parse_args()

    if not args.preview and not args.execute:
        print("Usar --preview para ver los datos o --execute para insertar.")
        sys.exit(0)

    # ── discover files ────────────────────────────────────────────────────────
    if args.file:
        xlsx_files = [args.file.resolve()]
    else:
        xlsx_files = discover_excel_files()

    if not xlsx_files:
        print(f"No se encontraron archivos .xlsx en {STORAGE_DIR}")
        print("Coloca el archivo Excel en la carpeta correspondiente dentro de storage/test_microcurriculos/")
        sys.exit(1)

    print(f"\n{'='*60}")
    print(f"Archivos .xlsx encontrados: {len(xlsx_files)}")
    for f in xlsx_files:
        print(f"  • {f.relative_to(REPO_ROOT)}")

    # ── parse all files ───────────────────────────────────────────────────────
    all_records: list[dict] = []
    for xlsx_path in xlsx_files:
        print(f"\n[PARSE] {xlsx_path.name}")
        records = parse_excel(xlsx_path)
        print(f"  → {len(records)} asignatura(s) extraídas")
        all_records.extend(records)

    if not all_records:
        print("\nNo se extrajeron registros. Verifica la estructura del Excel.")
        sys.exit(1)

    # ── preview without DB (use EXPLICIT_ESP_MAP directly) ───────────────────
    esp_map_local: dict[str, int | None] = {
        rec["programa"]: infer_esp_id(rec["programa"], [])
        for rec in all_records
    }
    print_preview(all_records, esp_map_local)

    if args.preview:
        print("[PREVIEW] Ejecutar con --execute para insertar en DB.")
        return

    # ── connect DB for execute ───────────────────────────────────────────────
    print("\n[DB] Conectando a Railway…")
    try:
        conn = get_conn()
        esps = fetch_especializaciones(conn)
        print(f"  → {len(esps)} especializaciones en DB")
    except Exception as exc:
        print(f"  ERROR de conexión: {exc}")
        sys.exit(1)

    # Refine esp_id with DB data
    esp_map: dict[str, int | None] = {}
    for rec in all_records:
        if rec["programa"] not in esp_map:
            esp_map[rec["programa"]] = infer_esp_id(rec["programa"], esps)

    # ── execute ───────────────────────────────────────────────────────────────
    if not args.yes:
        ans = input(f"¿Insertar {len(all_records)} registros en DB? [s/N]: ").strip().lower()
        if ans not in ("s", "si", "sí", "y", "yes"):
            print("Cancelado.")
            conn.close()
            return

    existing_hashes = fetch_existing_hashes(conn)
    inserted_micro = 0
    inserted_skills = 0
    skipped = 0
    errors = 0

    for rec in all_records:
        if rec["document_hash"] in existing_hashes:
            print(f"  [SKIP] Ya existe: {rec['asignatura'][:50]}")
            skipped += 1
            continue

        esp_id = esp_map.get(rec["programa"])
        if esp_id is None:
            print(f"  [WARN] No se encontró especializacion_id para '{rec['programa']}' — insertando sin vínculo")

        try:
            micro_id = insert_record(conn, rec, esp_id)
            if micro_id:
                existing_hashes.add(rec["document_hash"])
                skill_rows = insert_skills(conn, micro_id, rec)
                conn.commit()
                inserted_micro += 1
                inserted_skills += skill_rows
                print(f"  [OK] id={micro_id} '{rec['asignatura'][:50]}' — {skill_rows} skills")
            else:
                skipped += 1
        except Exception as exc:
            conn.rollback()
            print(f"  [ERROR] {rec['asignatura'][:50]}: {exc}")
            errors += 1

    conn.close()

    print(f"\n{'='*60}")
    print(f"RESULTADO FINAL")
    print(f"  Insertados  : {inserted_micro} microcurrículos")
    print(f"  Skills      : {inserted_skills} filas en microcurriculo_skills")
    print(f"  Omitidos    : {skipped} (ya existían)")
    print(f"  Errores     : {errors}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
